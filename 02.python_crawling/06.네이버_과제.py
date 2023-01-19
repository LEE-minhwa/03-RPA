import requests
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.Workbook()

ws = wb.create_sheet('주식데이터')
ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가손익'
ws['G1'] = '수익률'

ws['A2'] = '삼성전자'
ws['A3'] = 'SK하이닉스'
ws['A4'] = '카카오'

ws['C2'] = 85000
ws['C3'] = 120000
ws['C4'] = 145000

ws['D2'] = 20
ws['D3'] = 15
ws['D4'] = 10

for i in range(2,5,1):
    ws[f'E{i}'] = ws[f'B{i}']-ws[f'D{i}']
    ws[f'F{i}'] = (ws[f'B{i}']-ws[f'C{i}'])*ws[f'D{i}']
    ws[f'G{i}'] = ws[f'F{i}']/ws[f'C{i}']
    
wb.save('참가자_data.xlsx')

